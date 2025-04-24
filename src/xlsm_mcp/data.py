"""
Módulo para operaciones de lectura y escritura de datos en archivos Excel con macros (.xlsm)
"""

import os
import logging
from typing import Any, Dict, List, Optional, Union
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from xlsm_mcp.exceptions import DataError
from xlsm_mcp.workbook import open_workbook

logger = logging.getLogger("xlsm-mcp")

def parse_cell_range(start_cell: str, end_cell: Optional[str] = None) -> tuple:
    """
    Analiza referencias de celdas y devuelve coordenadas de fila/columna.
    
    Args:
        start_cell: Referencia de celda inicial (ej. "A1")
        end_cell: Referencia de celda final opcional (ej. "C5")
        
    Returns:
        Tupla (fila_inicio, columna_inicio, fila_fin, columna_fin)
        
    Raises:
        ValueError: Si las referencias de celda son inválidas
    """
    def _parse_cell(cell_ref: str) -> tuple:
        # Extrae letras de columna y número de fila
        import re
        match = re.match(r'^([A-Za-z]+)(\d+)$', cell_ref)
        if not match:
            raise ValueError(f"Referencia de celda inválida: {cell_ref}")
            
        col_str, row_str = match.groups()
        row = int(row_str)
        
        # Convierte letras de columna a número
        col = 0
        for char in col_str.upper():
            col = col * 26 + (ord(char) - ord('A') + 1)
            
        return row, col
    
    try:
        start_row, start_col = _parse_cell(start_cell)
        
        if end_cell:
            end_row, end_col = _parse_cell(end_cell)
        else:
            end_row, end_col = None, None
            
        return start_row, start_col, end_row, end_col
    except Exception as e:
        raise ValueError(f"Error al analizar rango de celdas: {str(e)}")

def read_excel_range(
    filepath: Union[str, Path],
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    include_formulas: bool = False
) -> List[Dict[str, Any]]:
    """
    Lee datos de un rango de celdas en Excel.
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        start_cell: Celda inicial (por defecto "A1")
        end_cell: Celda final (opcional)
        include_formulas: Si es True, incluye las fórmulas en lugar de sus valores
        
    Returns:
        Lista de diccionarios con los datos leídos
        
    Raises:
        DataError: Si ocurre un error al leer los datos
    """
    try:
        # Abrir el libro
        wb = open_workbook(filepath, read_only=True)
        
        if sheet_name not in wb.sheetnames:
            raise DataError(f"La hoja '{sheet_name}' no existe")
            
        ws = wb[sheet_name]

        # Analizar celda inicial
        if ':' in start_cell:
            start_cell, end_cell = start_cell.split(':')
            
        # Obtener coordenadas iniciales
        try:
            start_coords = parse_cell_range(start_cell)
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Formato de celda inicial inválido: {str(e)}")

        # Determinar coordenadas finales
        if end_cell:
            try:
                end_coords = parse_cell_range(end_cell)
                end_row, end_col = end_coords[0], end_coords[1]
            except ValueError as e:
                raise DataError(f"Formato de celda final inválido: {str(e)}")
        else:
            # Para una sola celda, usar las mismas coordenadas
            end_row, end_col = start_row, start_col

        # Validar límites del rango
        if start_row > ws.max_row or start_col > ws.max_column:
            raise DataError(
                f"Celda inicial fuera de límites. Dimensiones de la hoja: "
                f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            )

        data = []
        # Si es una sola celda o una sola fila, leer los valores directamente
        if start_row == end_row:
            row_data = {}
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=start_row, column=col)
                col_name = f"Columna_{get_column_letter(col)}"
                
                # Incluir fórmula o valor según configuración
                if include_formulas and cell.data_type == 'f':
                    row_data[col_name] = cell.value
                else:
                    row_data[col_name] = cell.value
                    
            if any(v is not None for v in row_data.values()):
                data.append(row_data)
        else:
            # Múltiples filas - usar fila de encabezado
            headers = []
            for col in range(start_col, end_col + 1):
                cell_value = ws.cell(row=start_row, column=col).value
                headers.append(str(cell_value) if cell_value is not None else f"Columna_{get_column_letter(col)}")

            # Obtener filas de datos
            for row in range(start_row + 1, end_row + 1):
                row_data = {}
                for col, header in enumerate(headers, start=start_col):
                    cell = ws.cell(row=row, column=col)
                    
                    # Incluir fórmula o valor según configuración
                    if include_formulas and cell.data_type == 'f':
                        row_data[header] = cell.value
                    else:
                        row_data[header] = cell.value
                        
                if any(v is not None for v in row_data.values()):
                    data.append(row_data)

        wb.close()
        return data
        
    except DataError:
        raise
    except Exception as e:
        logger.error(f"Error al leer rango de Excel: {e}")
        raise DataError(f"Error al leer datos: {str(e)}")

def write_data(
    filepath: str,
    sheet_name: str,
    data: List[Dict[str, Any]],
    start_cell: str = "A1"
) -> Dict[str, Any]:
    """
    Escribe datos en una hoja de Excel.
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        data: Lista de diccionarios con los datos a escribir
        start_cell: Celda inicial donde comenzar a escribir
        
    Returns:
        Diccionario con el resultado de la operación
        
    Raises:
        DataError: Si ocurre un error al escribir los datos
    """
    try:
        if not data:
            raise DataError("No se proporcionaron datos para escribir")
            
        # Abrir el libro
        wb = open_workbook(filepath, read_only=False)

        # Si la hoja no existe, crearla
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

        ws = wb[sheet_name]

        # Validar celda inicial
        try:
            start_coords = parse_cell_range(start_cell)
            start_row, start_col = start_coords[0], start_coords[1]
        except ValueError as e:
            raise DataError(f"Formato de celda inicial inválido: {str(e)}")

        # Escribir encabezados si hay datos
        if len(data) > 0:
            headers = list(data[0].keys())
            
            # Escribir encabezados
            for i, header in enumerate(headers):
                col = start_col + i
                cell = ws.cell(row=start_row, column=col)
                cell.value = header
                
            # Escribir datos
            for i, row_data in enumerate(data):
                row = start_row + i + 1
                for j, header in enumerate(headers):
                    col = start_col + j
                    cell = ws.cell(row=row, column=col)
                    cell.value = row_data.get(header)

        # Guardar y cerrar
        wb.save(filepath)
        wb.close()

        return {
            "message": f"Datos escritos en {sheet_name}",
            "rows_written": len(data)
        }
        
    except DataError:
        raise
    except Exception as e:
        logger.error(f"Error al escribir datos: {e}")
        raise DataError(f"Error al escribir datos: {str(e)}")

def append_data(
    filepath: str,
    sheet_name: str,
    data: List[Dict[str, Any]]
) -> Dict[str, Any]:
    """
    Añade datos al final de una hoja de Excel existente.
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        data: Lista de diccionarios con los datos a añadir
        
    Returns:
        Diccionario con el resultado de la operación
        
    Raises:
        DataError: Si ocurre un error al añadir los datos
    """
    try:
        if not data:
            raise DataError("No se proporcionaron datos para añadir")
            
        # Abrir el libro
        wb = open_workbook(filepath, read_only=False)

        # Si la hoja no existe, crearla
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(sheet_name)

        ws = wb[sheet_name]
        
        # Encontrar la última fila con datos
        last_row = 1
        for row in range(1, ws.max_row + 1):
            if any(ws.cell(row=row, column=col).value is not None for col in range(1, ws.max_column + 1)):
                last_row = row
                
        # Si la hoja está vacía, escribir encabezados
        if last_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
            # Escribir encabezados
            headers = list(data[0].keys())
            for i, header in enumerate(headers):
                col = i + 1
                cell = ws.cell(row=1, column=col)
                cell.value = header
                
            start_row = 2  # Comenzar a escribir datos en la fila 2
        else:
            # Verificar que los encabezados coinciden
            existing_headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
            data_headers = list(data[0].keys())
            
            # Si los encabezados no coinciden, mostrar advertencia
            if set(existing_headers) != set(data_headers):
                logger.warning("Los encabezados de los nuevos datos no coinciden con los existentes")
            
            start_row = last_row + 1
            
        # Escribir datos
        for i, row_data in enumerate(data):
            row = start_row + i
            for j, header in enumerate(row_data.keys()):
                col = j + 1
                cell = ws.cell(row=row, column=col)
                cell.value = row_data.get(header)

        # Guardar y cerrar
        wb.save(filepath)
        wb.close()

        return {
            "message": f"Datos añadidos en {sheet_name}",
            "rows_added": len(data),
            "start_row": start_row
        }
        
    except DataError:
        raise
    except Exception as e:
        logger.error(f"Error al añadir datos: {e}")
        raise DataError(f"Error al añadir datos: {str(e)}")