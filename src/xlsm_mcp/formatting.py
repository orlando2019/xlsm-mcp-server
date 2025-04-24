"""
Módulo para aplicar formato a celdas en archivos Excel con macros (.xlsm)

Este módulo proporciona funciones para aplicar formatos a celdas y rangos en hojas 
de Excel, incluyendo formatos condicionales, estilos de celda, y otras características
visuales.
"""

import logging
from typing import Any, Dict, List, Optional, Union
from pathlib import Path
import re
import os

from openpyxl.styles import (
    PatternFill, Border, Side, Alignment, Protection, Font,
    Color, NamedStyle, Fill
)
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule,
    FormulaRule, CellIsRule
)
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles.differential import DifferentialStyle

from xlsm_mcp.exceptions import ValidationError, FormattingError
from xlsm_mcp.workbook import open_workbook
from xlsm_mcp.validation import (
    validate_file_path, validate_sheet_name, 
    validate_cell_reference, validate_cell_range,
    validate_color
)

logger = logging.getLogger("xlsm-mcp")

def format_range(
    filepath: Union[str, Path],
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """
    Aplica formato a un rango de celdas.
    
    Esta función maneja todas las operaciones de formato de Excel incluyendo:
    - Propiedades de fuente (negrita, cursiva, tamaño, color, etc.)
    - Color de fondo/relleno de celda
    - Bordes (estilo y color)
    - Formato numérico
    - Alineación y ajuste de texto
    - Combinación de celdas
    - Protección
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        start_cell: Celda inicial
        end_cell: Celda final (opcional)
        bold: Si aplicar negrita
        italic: Si aplicar cursiva
        underline: Si aplicar subrayado
        font_size: Tamaño de fuente en puntos
        font_color: Color de fuente (código hexadecimal)
        bg_color: Color de fondo (código hexadecimal)
        border_style: Estilo de borde (thin, medium, thick, double)
        border_color: Color de borde (código hexadecimal)
        number_format: Cadena de formato numérico de Excel
        alignment: Alineación de texto (left, center, right, justify)
        wrap_text: Si ajustar texto
        merge_cells: Si combinar el rango
        protection: Configuración de protección de celdas
        
    Returns:
        Diccionario con estado de la operación
        
    Raises:
        ValidationError: Si los valores proporcionados no son válidos
        FormattingError: Si ocurre un error al aplicar el formato
        
    Ejemplos:
        ```python
        # Aplicar formato básico a una celda
        format_range("tabla.xlsx", "Hoja1", "A1", bold=True, font_size=14)
        
        # Formatear un rango de celdas con borde y relleno
        format_range(
            "informe.xlsm", "Datos", "B2", "D5", 
            bg_color="#FFFF00", border_style="thin", 
            border_color="#000000"
        )
        
        # Combinar celdas para un título
        format_range(
            "reporte.xlsx", "Portada", "A1", "E1",
            bold=True, font_size=16, alignment="center",
            merge_cells=True
        )
        ```
    """
    try:
        # Validar ruta de archivo
        file_path = validate_file_path(filepath, must_exist=True, 
                                      file_extensions=['.xlsx', '.xlsm'])
            
        # Abrir el libro
        wb = open_workbook(file_path, read_only=False)
        
        # Validar nombre de hoja
        validate_sheet_name(sheet_name)
        
        # Verificar que la hoja existe
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"La hoja '{sheet_name}' no existe")
            
        sheet = wb[sheet_name]
        
        # Validar referencias de celda
        start, end = validate_cell_range(start_cell, end_cell)
        
        # Si no se especificó celda final, usar la inicial
        if end is None:
            end = start
            
        # Convertir referencias de celda a coordenadas
        start_coord = sheet[start].coordinate
        end_coord = sheet[end].coordinate
        
        # Definir el rango de celdas
        cell_range = f"{start_coord}:{end_coord}"
            
        # Aplicar formato de fuente
        font_args = {
            "bold": bold,
            "italic": italic,
            "underline": 'single' if underline else None,
        }
        if font_size is not None:
            if not isinstance(font_size, int) or font_size <= 0:
                raise ValidationError(f"Tamaño de fuente inválido: {font_size}")
            font_args["size"] = font_size
            
        if font_color is not None:
            try:
                font_color = validate_color(font_color)
                # Convertir a formato RGB sin #
                rgb = font_color[1:] 
                # Asegurar que el color tiene prefijo FF para opacidad completa
                rgb = rgb if len(rgb) == 8 else f'FF{rgb}'
                font_args["color"] = Color(rgb=rgb)
            except ValidationError as e:
                raise
            except Exception as e:
                raise FormattingError(f"Color de fuente inválido: {str(e)}")
                
        font = Font(**font_args)
        
        # Aplicar relleno
        fill = None
        if bg_color is not None:
            try:
                bg_color = validate_color(bg_color)
                # Convertir a formato RGB sin #
                rgb = bg_color[1:]
                # Asegurar que el color tiene prefijo FF para opacidad completa
                rgb = rgb if len(rgb) == 8 else f'FF{rgb}'
                fill = PatternFill(
                    start_color=Color(rgb=rgb),
                    end_color=Color(rgb=rgb),
                    fill_type='solid'
                )
            except ValidationError as e:
                raise
            except Exception as e:
                raise FormattingError(f"Color de fondo inválido: {str(e)}")
        
        # Aplicar bordes
        border = None
        if border_style is not None:
            valid_styles = {'thin', 'medium', 'thick', 'double', 'dashed', 'dotted'}
            if border_style not in valid_styles:
                raise ValidationError(
                    f"Estilo de borde inválido: {border_style}. "
                    f"Valores válidos: {', '.join(valid_styles)}"
                )
            
            try:
                border_color_hex = "#000000"  # Negro por defecto
                if border_color:
                    border_color_hex = validate_color(border_color)
                    rgb = border_color_hex[1:]
                    rgb = rgb if len(rgb) == 8 else f'FF{rgb}'
                else:
                    rgb = "FF000000"  # Negro opaco
                    
                side = Side(style=border_style, color=Color(rgb=rgb))
                border = Border(left=side, right=side, top=side, bottom=side)
            except ValidationError as e:
                raise
            except Exception as e:
                raise FormattingError(f"Error al definir borde: {str(e)}")
        
        # Aplicar alineación
        alignment_obj = None
        if alignment is not None or wrap_text:
            align_args = {"wrap_text": wrap_text}
            if alignment is not None:
                valid_alignments = {'left', 'center', 'right', 'justify', 'general'}
                if alignment not in valid_alignments:
                    raise ValidationError(
                        f"Alineación inválida: {alignment}. "
                        f"Valores válidos: {', '.join(valid_alignments)}"
                    )
                align_args["horizontal"] = alignment
            alignment_obj = Alignment(**align_args)
            
        # Aplicar formato numérico
        number_format_str = number_format
        
        # Aplicar protección
        protection_obj = None
        if protection is not None:
            protection_args = {}
            if "locked" in protection:
                protection_args["locked"] = bool(protection["locked"])
            if "hidden" in protection:
                protection_args["hidden"] = bool(protection["hidden"])
            protection_obj = Protection(**protection_args)
        
        # Aplicar formatos a todas las celdas del rango
        for row in sheet[cell_range]:
            for cell in row:
                if font is not None:
                    cell.font = font
                if fill is not None:
                    cell.fill = fill
                if border is not None:
                    cell.border = border
                if alignment_obj is not None:
                    cell.alignment = alignment_obj
                if number_format_str is not None:
                    cell.number_format = number_format_str
                if protection_obj is not None:
                    cell.protection = protection_obj
        
        # Combinar celdas si se solicita
        if merge_cells and start != end:
            sheet.merge_cells(cell_range)
        
        # Guardar cambios
        wb.save(file_path)
        
        return {
            "success": True,
            "message": f"Formato aplicado correctamente al rango {cell_range} en hoja '{sheet_name}'"
        }
    except (ValidationError, FormattingError) as e:
        logger.error(f"Error al aplicar formato: {e}")
        return {
            "success": False,
            "error": str(e)
        }
    except Exception as e:
        logger.error(f"Error inesperado al aplicar formato: {e}")
        return {
            "success": False,
            "error": f"Error al aplicar formato: {str(e)}"
        }

def create_named_style(
    filepath: Union[str, Path],
    style_name: str,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False
) -> Dict[str, Any]:
    """
    Crea un estilo con nombre que puede ser reutilizado en el libro.
    
    Args:
        filepath: Ruta al archivo Excel
        style_name: Nombre único para el estilo
        [Resto de parámetros son iguales a format_range]
        
    Returns:
        Diccionario con estado de la operación
        
    Ejemplo:
        ```python
        # Crear un estilo para títulos
        create_named_style(
            "informe.xlsm", "Titulo",
            bold=True, font_size=14, 
            bg_color="#EEEEEE", border_style="thin"
        )
        
        # Aplicar el estilo a celdas
        apply_named_style("informe.xlsm", "Hoja1", "A1:A10", "Titulo")
        ```
    """
    try:
        # Validar ruta de archivo
        file_path = validate_file_path(filepath, must_exist=True, 
                                     file_extensions=['.xlsx', '.xlsm'])
        
        # Validar nombre de estilo
        if not style_name or not isinstance(style_name, str):
            raise ValidationError("El nombre del estilo no puede estar vacío")
            
        # Abrir el libro
        wb = open_workbook(file_path, read_only=False)
        
        # Crear el estilo
        style = NamedStyle(name=style_name)
        
        # Aplicar formato de fuente
        font_args = {
            "bold": bold,
            "italic": italic,
            "underline": 'single' if underline else None,
        }
        if font_size is not None:
            if not isinstance(font_size, int) or font_size <= 0:
                raise ValidationError(f"Tamaño de fuente inválido: {font_size}")
            font_args["size"] = font_size
            
        if font_color is not None:
            try:
                font_color = validate_color(font_color)
                # Convertir a formato RGB sin #
                rgb = font_color[1:] 
                # Asegurar que el color tiene prefijo FF para opacidad completa
                rgb = rgb if len(rgb) == 8 else f'FF{rgb}'
                font_args["color"] = Color(rgb=rgb)
            except ValidationError as e:
                raise
            except Exception as e:
                raise FormattingError(f"Color de fuente inválido: {str(e)}")
                
        style.font = Font(**font_args)
        
        # Aplicar relleno
        if bg_color is not None:
            try:
                bg_color = validate_color(bg_color)
                # Convertir a formato RGB sin #
                rgb = bg_color[1:]
                # Asegurar que el color tiene prefijo FF para opacidad completa
                rgb = rgb if len(rgb) == 8 else f'FF{rgb}'
                style.fill = PatternFill(
                    start_color=Color(rgb=rgb),
                    end_color=Color(rgb=rgb),
                    fill_type='solid'
                )
            except ValidationError as e:
                raise
            except Exception as e:
                raise FormattingError(f"Color de fondo inválido: {str(e)}")
        
        # Aplicar bordes
        if border_style is not None:
            valid_styles = {'thin', 'medium', 'thick', 'double', 'dashed', 'dotted'}
            if border_style not in valid_styles:
                raise ValidationError(
                    f"Estilo de borde inválido: {border_style}. "
                    f"Valores válidos: {', '.join(valid_styles)}"
                )
            
            try:
                border_color_hex = "#000000"  # Negro por defecto
                if border_color:
                    border_color_hex = validate_color(border_color)
                    rgb = border_color_hex[1:]
                    rgb = rgb if len(rgb) == 8 else f'FF{rgb}'
                else:
                    rgb = "FF000000"  # Negro opaco
                    
                side = Side(style=border_style, color=Color(rgb=rgb))
                style.border = Border(left=side, right=side, top=side, bottom=side)
            except ValidationError as e:
                raise
            except Exception as e:
                raise FormattingError(f"Error al definir borde: {str(e)}")
        
        # Aplicar alineación
        if alignment is not None or wrap_text:
            align_args = {"wrap_text": wrap_text}
            if alignment is not None:
                valid_alignments = {'left', 'center', 'right', 'justify', 'general'}
                if alignment not in valid_alignments:
                    raise ValidationError(
                        f"Alineación inválida: {alignment}. "
                        f"Valores válidos: {', '.join(valid_alignments)}"
                    )
                align_args["horizontal"] = alignment
            style.alignment = Alignment(**align_args)
            
        # Aplicar formato numérico
        if number_format is not None:
            style.number_format = number_format
            
        # Añadir el estilo al libro
        if style_name in wb.named_styles:
            # Sobrescribir estilo existente
            for style_idx, style_obj in enumerate(wb._named_styles):
                if style_obj.name == style_name:
                    wb._named_styles[style_idx] = style
        else:
            # Añadir nuevo estilo
            wb.add_named_style(style)
        
        # Guardar cambios
        wb.save(file_path)
        
        return {
            "success": True,
            "message": f"Estilo '{style_name}' creado correctamente"
        }
    except (ValidationError, FormattingError) as e:
        logger.error(f"Error al crear estilo: {e}")
        return {
            "success": False,
            "error": str(e)
        }
    except Exception as e:
        logger.error(f"Error inesperado al crear estilo: {e}")
        return {
            "success": False,
            "error": f"Error al crear estilo: {str(e)}"
        }

def apply_named_style(
    filepath: Union[str, Path],
    sheet_name: str,
    cell_range: str,
    style_name: str
) -> Dict[str, Any]:
    """
    Aplica un estilo con nombre a un rango de celdas.
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        cell_range: Rango de celdas (ej. "A1:B10")
        style_name: Nombre del estilo a aplicar
        
    Returns:
        Diccionario con estado de la operación
    """
    try:
        # Validar ruta de archivo
        file_path = validate_file_path(filepath, must_exist=True, 
                                     file_extensions=['.xlsx', '.xlsm'])
        
        # Validar nombre de hoja
        validate_sheet_name(sheet_name)
        
        # Abrir el libro
        wb = open_workbook(file_path, read_only=False)
        
        # Verificar que la hoja existe
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"La hoja '{sheet_name}' no existe")
            
        sheet = wb[sheet_name]
        
        # Verificar que el estilo existe
        if style_name not in wb.named_styles:
            raise ValidationError(f"El estilo '{style_name}' no existe en el libro")
        
        # Aplicar estilo a las celdas
        for row in sheet[cell_range]:
            for cell in row:
                cell.style = style_name
        
        # Guardar cambios
        wb.save(file_path)
        
        return {
            "success": True,
            "message": f"Estilo '{style_name}' aplicado al rango {cell_range} en hoja '{sheet_name}'"
        }
    except (ValidationError, FormattingError) as e:
        logger.error(f"Error al aplicar estilo: {e}")
        return {
            "success": False,
            "error": str(e)
        }
    except Exception as e:
        logger.error(f"Error inesperado al aplicar estilo: {e}")
        return {
            "success": False,
            "error": f"Error al aplicar estilo: {str(e)}"
        }

def apply_conditional_formatting(
    filepath: str,
    sheet_name: str,
    range_string: str,
    rule_type: str,
    formula: Optional[str] = None,
    color_scale: Optional[list] = None,
    data_bar: Optional[Dict] = None,
    icon_set: Optional[Dict] = None,
    operator: Optional[str] = None,
    value: Optional[Any] = None,
    text: Optional[str] = None
) -> Dict[str, Any]:
    """
    Aplica formato condicional a un rango de celdas.
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        range_string: Cadena de rango (ej. "A1:C10")
        rule_type: Tipo de regla (formula, color_scale, data_bar, icon_set, cell_is, contains_text)
        formula: Fórmula para regla basada en fórmula
        color_scale: Lista de colores para escala de color
        data_bar: Configuración para barra de datos
        icon_set: Configuración para conjunto de iconos
        operator: Operador para regla "cell_is" (equal, not_equal, greater_than, etc.)
        value: Valor para regla "cell_is"
        text: Texto para regla "contains_text"
        
    Returns:
        Diccionario con estado de la operación
        
    Raises:
        ValidationError: Si los valores proporcionados no son válidos
        FormattingError: Si ocurre un error al aplicar el formato condicional
    """
    try:
        # Validar tipo de regla
        valid_rule_types = ["formula", "color_scale", "data_bar", "icon_set", "cell_is", "contains_text"]
        if rule_type not in valid_rule_types:
            raise ValidationError(f"Tipo de regla inválido. Debe ser uno de: {', '.join(valid_rule_types)}")
            
        # Abrir el libro
        wb = open_workbook(filepath, read_only=False)
        
        # Verificar que la hoja existe
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"La hoja '{sheet_name}' no existe")
            
        ws = wb[sheet_name]
        
        # Crear la regla según el tipo
        if rule_type == "formula":
            if not formula:
                raise ValidationError("Se requiere una fórmula para el tipo de regla 'formula'")
            rule = FormulaRule(
                formula=[formula],
                stopIfTrue=False,
                fill=PatternFill(start_color="FFFF00", end_color="FFFF00")  # Amarillo por defecto
            )
        elif rule_type == "color_scale":
            if not color_scale or len(color_scale) < 2:
                color_scale = ["FFFF0000", "FFFFFF00", "FF00FF00"]  # Rojo a amarillo a verde
            rule = ColorScaleRule(
                start_type='min',
                start_color=color_scale[0],
                end_type='max',
                end_color=color_scale[-1],
                mid_type='percentile' if len(color_scale) > 2 else None,
                mid_color=color_scale[1] if len(color_scale) > 2 else None
            )
        elif rule_type == "data_bar":
            color = "FF638EC6"  # Azul por defecto
            if data_bar and "color" in data_bar:
                color = data_bar["color"]
            rule = DataBarRule(
                start_type='min',
                end_type='max',
                color=color
            )
        elif rule_type == "icon_set":
            icon_style = "3Arrows"  # Flechas por defecto
            if icon_set and "style" in icon_set:
                icon_style = icon_set["style"]
            rule = IconSetRule(
                icon_style=icon_style
            )
        elif rule_type == "cell_is":
            if operator is None:
                raise ValidationError("Se requiere un operador para el tipo de regla 'cell_is'")
            if value is None:
                raise ValidationError("Se requiere un valor para el tipo de regla 'cell_is'")
                
            rule = CellIsRule(
                operator=operator,
                formula=[str(value)],
                stopIfTrue=False,
                fill=PatternFill(start_color="FFFF00", end_color="FFFF00")  # Amarillo por defecto
            )
        elif rule_type == "contains_text":
            if text is None:
                raise ValidationError("Se requiere texto para el tipo de regla 'contains_text'")
                
            rule = FormulaRule(
                formula=[f'NOT(ISERROR(SEARCH("{text}",A1)))'],
                stopIfTrue=False,
                fill=PatternFill(start_color="FFFF00", end_color="FFFF00")  # Amarillo por defecto
            )
            
        # Aplicar la regla al rango
        ws.conditional_formatting.add(range_string, rule)
        
        # Guardar y cerrar el libro
        wb.save(filepath)
        wb.close()
        
        return {
            "success": True,
            "message": f"Formato condicional aplicado correctamente al rango {range_string}"
        }
    except ValidationError as e:
        logger.error(str(e))
        raise
    except FormattingError as e:
        logger.error(str(e))
        raise
    except Exception as e:
        logger.error(f"Error al aplicar formato condicional: {e}")
        raise FormattingError(f"Error al aplicar formato condicional: {str(e)}")

def clear_formatting(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None
) -> Dict[str, Any]:
    """
    Elimina todo el formato de un rango de celdas.
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        start_cell: Celda inicial
        end_cell: Celda final (opcional)
        
    Returns:
        Diccionario con estado de la operación
        
    Raises:
        ValidationError: Si los valores proporcionados no son válidos
        FormattingError: Si ocurre un error al eliminar el formato
    """
    try:
        # Validar ruta de archivo
        file_path = validate_file_path(filepath, must_exist=True, 
                                      file_extensions=['.xlsx', '.xlsm'])
        
        # Validar nombre de hoja
        validate_sheet_name(sheet_name)
        
        # Abrir el libro
        wb = open_workbook(file_path, read_only=False)
        
        # Verificar que la hoja existe
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"La hoja '{sheet_name}' no existe")
            
        ws = wb[sheet_name]
        
        # Validar referencias de celda
        start, end = validate_cell_range(start_cell, end_cell)
        
        # Si no se especificó celda final, usar la inicial
        if end is None:
            end = start
            
        # Convertir referencias de celda a coordenadas
        start_coord = ws[start].coordinate
        end_coord = ws[end].coordinate
        
        # Definir el rango de celdas
        cell_range = f"{start_coord}:{end_coord}"
        
        # Crear objetos de formato predeterminados
        default_font = Font()
        default_border = Border()
        default_fill = PatternFill()
        
        # Eliminar formato en cada celda del rango
        for row in ws[cell_range]:
            for cell in row:
                cell.font = default_font
                cell.border = default_border
                cell.fill = default_fill
                cell.number_format = "General"
                cell.alignment = Alignment()
                cell.protection = Protection()
                
        # Guardar cambios
        wb.save(file_path)
        
        return {
            "success": True,
            "message": f"Formato eliminado correctamente del rango {cell_range}"
        }
    except (ValidationError, FormattingError) as e:
        logger.error(f"Error al eliminar formato: {e}")
        return {
            "success": False,
            "error": str(e)
        }
    except Exception as e:
        logger.error(f"Error inesperado al eliminar formato: {e}")
        return {
            "success": False,
            "error": f"Error al eliminar formato: {str(e)}"
        }

def set_column_width(
    filepath: Union[str, Path],
    sheet_name: str,
    column: str,
    width: float
) -> Dict[str, Any]:
    """
    Establece el ancho de una columna.
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        column: Letra o referencia de columna (ej. "A", "BC")
        width: Ancho de columna en unidades de carácter
        
    Returns:
        Diccionario con estado de la operación
        
    Raises:
        ValidationError: Si los valores proporcionados no son válidos
        FormattingError: Si ocurre un error al establecer el ancho
    """
    try:
        # Validar ruta de archivo
        file_path = validate_file_path(filepath, must_exist=True, 
                                      file_extensions=['.xlsx', '.xlsm'])
        
        # Validar nombre de hoja
        validate_sheet_name(sheet_name)
        
        # Validar letra de columna
        if not re.match(r'^[A-Za-z]+$', column):
            raise ValidationError(f"Referencia de columna inválida: {column}")
            
        # Asegurar que el ancho es positivo
        if width <= 0:
            raise ValidationError("El ancho debe ser un valor positivo")
            
        # Abrir el libro
        wb = open_workbook(file_path, read_only=False)
        
        # Verificar que la hoja existe
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"La hoja '{sheet_name}' no existe")
            
        ws = wb[sheet_name]
        
        # Establecer ancho de columna
        ws.column_dimensions[column.upper()].width = width
        
        # Guardar cambios
        wb.save(file_path)
        
        return {
            "success": True,
            "message": f"Ancho de columna {column} establecido a {width} unidades"
        }
    except (ValidationError, FormattingError) as e:
        logger.error(f"Error al establecer ancho de columna: {e}")
        return {
            "success": False,
            "error": str(e)
        }
    except Exception as e:
        logger.error(f"Error inesperado al establecer ancho de columna: {e}")
        return {
            "success": False,
            "error": f"Error al establecer ancho de columna: {str(e)}"
        }

def set_row_height(
    filepath: Union[str, Path],
    sheet_name: str,
    row: int,
    height: float
) -> Dict[str, Any]:
    """
    Establece la altura de una fila.
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        row: Número de fila
        height: Altura de fila en puntos
        
    Returns:
        Diccionario con estado de la operación
        
    Raises:
        ValidationError: Si los valores proporcionados no son válidos
        FormattingError: Si ocurre un error al establecer la altura
    """
    try:
        # Validar ruta de archivo
        file_path = validate_file_path(filepath, must_exist=True, 
                                      file_extensions=['.xlsx', '.xlsm'])
        
        # Validar nombre de hoja
        validate_sheet_name(sheet_name)
        
        # Validar número de fila
        if row <= 0:
            raise ValidationError("El número de fila debe ser positivo")
            
        # Asegurar que la altura es positiva
        if height <= 0:
            raise ValidationError("La altura debe ser un valor positivo")
            
        # Abrir el libro
        wb = open_workbook(file_path, read_only=False)
        
        # Verificar que la hoja existe
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"La hoja '{sheet_name}' no existe")
            
        ws = wb[sheet_name]
        
        # Establecer altura de fila
        ws.row_dimensions[row].height = height
        
        # Guardar cambios
        wb.save(file_path)
        
        return {
            "success": True,
            "message": f"Altura de fila {row} establecida a {height} puntos"
        }
    except (ValidationError, FormattingError) as e:
        logger.error(f"Error al establecer altura de fila: {e}")
        return {
            "success": False,
            "error": str(e)
        }
    except Exception as e:
        logger.error(f"Error inesperado al establecer altura de fila: {e}")
        return {
            "success": False,
            "error": f"Error al establecer altura de fila: {str(e)}"
        }

def add_conditional_formatting(
    filepath: Union[str, Path],
    sheet_name: str,
    cell_range: str,
    rule_type: str,
    formula: str = None,
    operator: str = None,
    values: List[Any] = None,
    styles: Dict[str, Any] = None,
    colors: List[str] = None,
    icon_style: str = None,
    stopif_true: bool = False,
    priority: int = 1
) -> Dict[str, Any]:
    """
    Aplica formato condicional a un rango de celdas en un archivo Excel.
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la hoja donde aplicar el formato
        cell_range: Rango de celdas para aplicar el formato (ej. "A1:B10")
        rule_type: Tipo de regla de formato condicional:
                  'cell_is', 'formula', 'color_scale', 'data_bar', 'icon_set'
        formula: Fórmula para la regla de tipo 'formula'
        operator: Operador para regla 'cell_is' ('between', 'not between', '>', '<', '>=', '<=', '=', '!=')
        values: Valores para la regla (depende del tipo de regla)
        styles: Estilos para aplicar (diccionario con propiedades de estilo)
        colors: Lista de colores para reglas de tipo 'color_scale' o 'data_bar'
        icon_style: Estilo de iconos para reglas de tipo 'icon_set' ('3_traffic_lights', '3_symbols', etc.)
        stopif_true: Si True, no se evaluarán más reglas después de esta si se cumple
        priority: Prioridad de la regla (menor número = mayor prioridad)
        
    Returns:
        Dict: Resultado de la operación con claves 'success' y 'message' o 'error'
        
    Raises:
        ValidationError: Si algún parámetro no es válido
        FormattingError: Si hay un error al aplicar el formato condicional
    """
    try:
        # Validar ruta del archivo
        validate_file_path(filepath)
        
        # Validar nombre de hoja
        validate_sheet_name(sheet_name)
        
        # Validar rango de celdas
        validate_cell_range(cell_range)
        
        # Validar tipo de regla
        valid_rule_types = ['cell_is', 'formula', 'color_scale', 'data_bar', 'icon_set']
        if rule_type not in valid_rule_types:
            valid_types_str = ", ".join([f"'{t}'" for t in valid_rule_types])
            raise ValidationError(f"Tipo de regla inválido: '{rule_type}'. Debe ser uno de: {valid_types_str}")
        
        # Validar operador para 'cell_is'
        if rule_type == 'cell_is' and operator:
            valid_operators = ['between', 'not between', '>', '<', '>=', '<=', '=', '!=']
            if operator not in valid_operators:
                valid_ops_str = ", ".join([f"'{op}'" for op in valid_operators])
                raise ValidationError(f"Operador inválido: '{operator}'. Debe ser uno de: {valid_ops_str}")
        
        # Abrir el libro y obtener la hoja
        wb = open_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"La hoja '{sheet_name}' no existe en el archivo")
        
        ws = wb[sheet_name]
        
        # Crear la regla de formato condicional según el tipo
        rule = None
        
        if rule_type == 'cell_is':
            if not operator:
                raise ValidationError("Se requiere un operador para el tipo de regla 'cell_is'")
            if not values:
                raise ValidationError("Se requieren valores para el tipo de regla 'cell_is'")
            
            # Crear estilo diferencial
            dxf = None
            if styles:
                dxf = _create_differential_style(styles)
            
            rule = CellIsRule(
                operator=operator,
                formula=values if isinstance(values, list) else [values],
                stopIfTrue=stopif_true,
                dxf=dxf
            )
            
        elif rule_type == 'formula':
            if not formula:
                raise ValidationError("Se requiere una fórmula para el tipo de regla 'formula'")
            
            # Crear estilo diferencial
            dxf = None
            if styles:
                dxf = _create_differential_style(styles)
            
            rule = FormulaRule(
                formula=[formula],
                stopIfTrue=stopif_true,
                dxf=dxf
            )
            
        elif rule_type == 'color_scale':
            if not colors or len(colors) < 2:
                raise ValidationError("Se requieren al menos 2 colores para el tipo de regla 'color_scale'")
            
            # Valores por defecto si no se proporcionan
            if not values:
                if len(colors) == 2:
                    values = ['min', 'max']
                elif len(colors) == 3:
                    values = ['min', '50', 'max']
                else:
                    raise ValidationError("Para 'color_scale' con más de 3 colores, se deben proporcionar valores explícitamente")
            
            if len(values) != len(colors):
                raise ValidationError(f"La cantidad de valores ({len(values)}) debe coincidir con la cantidad de colores ({len(colors)})")
            
            rule = ColorScaleRule(
                start_type=values[0] if isinstance(values[0], str) else 'num',
                start_value=None if isinstance(values[0], str) else values[0],
                start_color=colors[0],
                end_type=values[-1] if isinstance(values[-1], str) else 'num',
                end_value=None if isinstance(values[-1], str) else values[-1],
                end_color=colors[-1],
                mid_type=values[1] if len(values) > 2 and isinstance(values[1], str) else None,
                mid_value=None if len(values) <= 2 or isinstance(values[1], str) else values[1],
                mid_color=colors[1] if len(colors) > 2 else None
            )
            
        elif rule_type == 'data_bar':
            if not colors:
                raise ValidationError("Se requiere al menos un color para el tipo de regla 'data_bar'")
            
            color = colors[0] if isinstance(colors, list) else colors
            
            rule = DataBarRule(
                start_type='min',
                start_value=None,
                end_type='max',
                end_value=None,
                color=color,
                showValue=True,
                minLength=None,
                maxLength=None
            )
            
        elif rule_type == 'icon_set':
            if not icon_style:
                raise ValidationError("Se requiere un estilo de iconos para el tipo de regla 'icon_set'")
            
            valid_icon_styles = [
                '3_arrows', '3_arrows_gray', '3_flags', '3_traffic_lights', '3_signs', '3_symbols', '3_symbols_2',
                '4_arrows', '4_arrows_gray', '4_ratings', '4_traffic_lights', '5_arrows', '5_arrows_gray', '5_ratings'
            ]
            
            if icon_style not in valid_icon_styles:
                raise ValidationError(f"Estilo de iconos inválido: '{icon_style}'. Debe ser uno de: {', '.join(valid_icon_styles)}")
            
            # Determinar cuántos valores necesitamos según el estilo
            icon_count = int(icon_style[0])
            
            # Valores por defecto si no se proporcionan
            if not values:
                if icon_count == 3:
                    values = [0, 33, 67]
                elif icon_count == 4:
                    values = [0, 25, 50, 75]
                elif icon_count == 5:
                    values = [0, 20, 40, 60, 80]
            
            if len(values) != icon_count:
                raise ValidationError(f"Se requieren {icon_count} valores para el estilo de iconos '{icon_style}'")
            
            rule = IconSetRule(
                icon_style=icon_style,
                type=['percent'] * icon_count,
                values=values,
                showValue=True,
                reverse=False
            )
        
        # Agregar la regla a la hoja
        ws.conditional_formatting.add(cell_range, rule)
        
        # Establecer prioridad si se especifica
        if priority is not None and priority > 0:
            for i, cf_rule in enumerate(ws.conditional_formatting._cf_rules[cell_range]):
                if cf_rule == rule:
                    cf_rule.priority = priority
                    break
        
        # Guardar el libro
        wb.save(filepath)
        
        return {
            "success": True,
            "message": f"Formato condicional aplicado correctamente a {cell_range} en la hoja '{sheet_name}'"
        }
        
    except ValidationError as ve:
        return {
            "success": False,
            "error": f"Error de validación: {str(ve)}"
        }
    except FormattingError as fe:
        return {
            "success": False,
            "error": f"Error de formato: {str(fe)}"
        }
    except Exception as e:
        return {
            "success": False,
            "error": f"Error al aplicar formato condicional: {str(e)}"
        }

def _create_differential_style(style_dict: Dict[str, Any]) -> DifferentialStyle:
    """
    Crea un estilo diferencial a partir de un diccionario de propiedades de estilo.
    
    Args:
        style_dict: Diccionario con propiedades de estilo
                   Claves soportadas: 'font', 'fill', 'border', 'alignment', 'number_format'
    
    Returns:
        DifferentialStyle: Objeto de estilo diferencial para usar en formatos condicionales
        
    Raises:
        FormattingError: Si hay un error al crear el estilo
    """
    try:
        font = None
        fill = None
        border = None
        alignment = None
        number_format = None
        
        # Procesar fuente
        if 'font' in style_dict:
            font_props = style_dict['font']
            font = Font(
                name=font_props.get('name'),
                size=font_props.get('size'),
                bold=font_props.get('bold'),
                italic=font_props.get('italic'),
                underline=font_props.get('underline'),
                strike=font_props.get('strike'),
                color=font_props.get('color')
            )
        
        # Procesar relleno
        if 'fill' in style_dict:
            fill_props = style_dict['fill']
            pattern_type = fill_props.get('pattern_type', 'solid')
            
            if pattern_type == 'solid':
                fg_color = fill_props.get('fg_color')
                bg_color = fill_props.get('bg_color')
                fill = PatternFill(
                    patternType=pattern_type,
                    fgColor=fg_color,
                    bgColor=bg_color
                )
            else:
                # Para otros tipos de relleno, simplemente usar un PatternFill básico
                # El soporte para GradientFill en DifferentialStyle es limitado
                fg_color = fill_props.get('fg_color')
                fill = PatternFill(
                    patternType='solid',
                    fgColor=fg_color
                )
        
        # Procesar bordes
        if 'border' in style_dict:
            border_props = style_dict['border']
            
            # Función auxiliar para crear un objeto Side
            def create_side(side_props):
                if not side_props:
                    return None
                return Side(
                    style=side_props.get('style'),
                    color=side_props.get('color')
                )
            
            border = Border(
                left=create_side(border_props.get('left')),
                right=create_side(border_props.get('right')),
                top=create_side(border_props.get('top')),
                bottom=create_side(border_props.get('bottom')),
                diagonal=create_side(border_props.get('diagonal')),
                diagonalUp=border_props.get('diagonal_up', False),
                diagonalDown=border_props.get('diagonal_down', False)
            )
        
        # Procesar alineación
        if 'alignment' in style_dict:
            align_props = style_dict['alignment']
            alignment = Alignment(
                horizontal=align_props.get('horizontal'),
                vertical=align_props.get('vertical'),
                textRotation=align_props.get('text_rotation'),
                wrapText=align_props.get('wrap_text'),
                shrinkToFit=align_props.get('shrink_to_fit'),
                indent=align_props.get('indent'),
                justifyLastLine=align_props.get('justify_last_line'),
                readingOrder=align_props.get('reading_order')
            )
        
        # Procesar formato numérico
        if 'number_format' in style_dict:
            number_format = style_dict['number_format']
        
        # Crear y devolver el estilo diferencial
        return DifferentialStyle(
            font=font,
            fill=fill,
            border=border,
            alignment=alignment,
            numFmt=number_format
        )
    except Exception as e:
        raise FormattingError(f"Error al crear estilo diferencial: {str(e)}")

def remove_conditional_formatting(
    filepath: Union[str, Path],
    sheet_name: str,
    cell_range: Optional[str] = None
) -> Dict[str, Any]:
    """
    Elimina formato condicional de un rango de celdas o de toda la hoja.
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        cell_range: Rango de celdas a limpiar (si es None, limpia toda la hoja)
        
    Returns:
        Diccionario con el estado de la operación
        
    Raises:
        ValidationError: Si los parámetros no son válidos
        FormattingError: Si hay un error al eliminar el formato condicional
    """
    try:
        # Validar ruta de archivo
        file_path = validate_file_path(filepath, must_exist=True, 
                                     file_extensions=['.xlsx', '.xlsm'])
        
        # Validar nombre de hoja
        validate_sheet_name(sheet_name)
        
        # Validar rango de celdas si se proporciona
        if cell_range:
            validate_cell_range(cell_range)
        
        # Abrir el libro
        wb = open_workbook(file_path, read_only=False)
        
        # Verificar que la hoja existe
        if sheet_name not in wb.sheetnames:
            raise ValidationError(f"La hoja '{sheet_name}' no existe")
            
        ws = wb[sheet_name]
        
        if cell_range:
            # Eliminar formato condicional solo del rango especificado
            to_remove = []
            for cf_range, rules in ws.conditional_formatting.items():
                if cf_range == cell_range:
                    to_remove.append(cf_range)
            
            for cf_range in to_remove:
                del ws.conditional_formatting[cf_range]
                
            message = f"Formato condicional eliminado del rango {cell_range}"
        else:
            # Eliminar todo el formato condicional de la hoja
            ws.conditional_formatting = {}
            message = f"Todo el formato condicional eliminado de la hoja '{sheet_name}'"
        
        # Guardar cambios
        wb.save(file_path)
        
        return {
            "success": True,
            "message": message
        }
    except (ValidationError, FormattingError) as e:
        logger.error(f"Error al eliminar formato condicional: {e}")
        return {
            "success": False,
            "error": str(e)
        }
    except Exception as e:
        logger.error(f"Error inesperado al eliminar formato condicional: {e}")
        return {
            "success": False,
            "error": f"Error al eliminar formato condicional: {str(e)}"
        }

def validate_sheet_name(sheet_name: str) -> None:
    """
    Valida el nombre de una hoja.
    
    Args:
        sheet_name: Nombre de la hoja a validar
        
    Raises:
        ValidationError: Si el nombre de la hoja no es válido
    """
    if not sheet_name or not isinstance(sheet_name, str):
        raise ValidationError("El nombre de la hoja no puede estar vacío y debe ser una cadena de texto")
    
    # Verificar longitud máxima (31 caracteres en Excel)
    if len(sheet_name) > 31:
        raise ValidationError("El nombre de la hoja no puede exceder los 31 caracteres")
    
    # Verificar caracteres inválidos: /, \, ?, *, [, ], :
    invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
    for char in invalid_chars:
        if char in sheet_name:
            raise ValidationError(f"El nombre de la hoja contiene un carácter inválido: '{char}'")

def validate_cell_range(cell_range: str) -> bool:
    """
    Valida que un rango de celdas tenga un formato válido (ej. "A1:B10" o "A1").
    
    Args:
        cell_range: Rango de celdas a validar
        
    Returns:
        bool: True si el rango es válido
        
    Raises:
        ValidationError: Si el rango no es válido
    """
    if not cell_range or not isinstance(cell_range, str):
        raise ValidationError("El rango de celdas debe ser una cadena no vacía")
    
    # Expresiones regulares para validar celdas
    single_cell_pattern = r'^[A-Za-z]{1,3}[1-9][0-9]{0,6}$'
    cell_range_pattern = r'^[A-Za-z]{1,3}[1-9][0-9]{0,6}:[A-Za-z]{1,3}[1-9][0-9]{0,6}$'
    
    # Verificar si es una celda individual o un rango
    if re.match(single_cell_pattern, cell_range):
        return True
    elif re.match(cell_range_pattern, cell_range):
        # Extraer las celdas inicial y final
        start_cell, end_cell = cell_range.split(':')
        
        # Extraer columnas y filas
        start_col = ''.join(filter(str.isalpha, start_cell))
        start_row = int(''.join(filter(str.isdigit, start_cell)))
        
        end_col = ''.join(filter(str.isalpha, end_cell))
        end_row = int(''.join(filter(str.isdigit, end_cell)))
        
        # Convertir columnas a índices numéricos (A=1, B=2, ...)
        start_col_index = 0
        for char in start_col.upper():
            start_col_index = start_col_index * 26 + (ord(char) - ord('A') + 1)
        
        end_col_index = 0
        for char in end_col.upper():
            end_col_index = end_col_index * 26 + (ord(char) - ord('A') + 1)
        
        # Verificar que la celda inicial sea menor que la final
        if (start_col_index > end_col_index) or (start_col_index == end_col_index and start_row > end_row):
            raise ValidationError(f"Rango de celdas inválido: {cell_range}. La celda inicial debe ser menor que la final")
        
        return True
    else:
        raise ValidationError(f"Formato de rango de celdas inválido: {cell_range}. Debe ser 'A1' o 'A1:B10'")