"""
Módulo para operaciones con hojas de Excel en archivos con macros (.xlsm)
"""

import logging
from typing import Any, Dict, List, Optional, Union
from copy import copy

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, Border, PatternFill, Side

from xlsm_mcp.exceptions import SheetError, ValidationError
from xlsm_mcp.workbook import open_workbook

logger = logging.getLogger("xlsm-mcp")

def parse_cell_range(start_cell: str, end_cell: Optional[str] = None) -> tuple:
    """
    Analiza referencias de celdas y devuelve coordenadas de fila/columna.
    
    Args:
        start_cell: Referencia de celda inicial (ej. "A1")
        end_cell: Referencia de celda final opcional (ej. "C5")
        
    Returns:
        Tupla (fila_inicio, columna_inicio, fila_fin, columna_fin)
        
    Raises:
        ValueError: Si las referencias de celda son inválidas
    """
    def _parse_cell(cell_ref: str) -> tuple:
        # Extrae letras de columna y número de fila
        import re
        match = re.match(r'^([A-Za-z]+)(\d+)$', cell_ref)
        if not match:
            raise ValueError(f"Referencia de celda inválida: {cell_ref}")
            
        col_str, row_str = match.groups()
        row = int(row_str)
        
        # Convierte letras de columna a número
        col = 0
        for char in col_str.upper():
            col = col * 26 + (ord(char) - ord('A') + 1)
            
        return row, col
    
    try:
        start_row, start_col = _parse_cell(start_cell)
        
        if end_cell:
            end_row, end_col = _parse_cell(end_cell)
        else:
            end_row, end_col = None, None
            
        return start_row, start_col, end_row, end_col
    except Exception as e:
        raise ValueError(f"Error al analizar rango de celdas: {str(e)}")

def create_worksheet(filepath: str, sheet_name: str) -> Dict[str, Any]:
    """
    Crea una nueva hoja en un libro de Excel existente.
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la nueva hoja
        
    Returns:
        Diccionario con el resultado de la operación
        
    Raises:
        SheetError: Si ocurre un error al crear la hoja
    """
    try:
        # Abrir el libro
        wb = open_workbook(filepath, read_only=False)
        
        # Verificar si la hoja ya existe
        if sheet_name in wb.sheetnames:
            raise SheetError(f"La hoja '{sheet_name}' ya existe")
            
        # Crear nueva hoja
        wb.create_sheet(sheet_name)
        
        # Guardar y cerrar
        wb.save(filepath)
        wb.close()
        
        return {"message": f"Hoja '{sheet_name}' creada correctamente"}
    except SheetError:
        raise
    except Exception as e:
        logger.error(f"Error al crear hoja: {e}")
        raise SheetError(f"Error al crear hoja: {str(e)}")

def copy_sheet(filepath: str, source_sheet: str, target_sheet: str) -> Dict[str, Any]:
    """
    Copia una hoja dentro del mismo libro.
    
    Args:
        filepath: Ruta al archivo Excel
        source_sheet: Nombre de la hoja de origen
        target_sheet: Nombre de la hoja de destino
        
    Returns:
        Diccionario con el resultado de la operación
        
    Raises:
        SheetError: Si ocurre un error al copiar la hoja
    """
    try:
        # Abrir el libro
        wb = open_workbook(filepath, read_only=False)
        
        # Verificar que la hoja de origen existe
        if source_sheet not in wb.sheetnames:
            raise SheetError(f"La hoja de origen '{source_sheet}' no existe")
            
        # Verificar que la hoja de destino no existe
        if target_sheet in wb.sheetnames:
            raise SheetError(f"La hoja de destino '{target_sheet}' ya existe")
            
        # Copiar hoja
        source = wb[source_sheet]
        target = wb.copy_worksheet(source)
        target.title = target_sheet
        
        # Guardar y cerrar
        wb.save(filepath)
        wb.close()
        
        return {"message": f"Hoja '{source_sheet}' copiada a '{target_sheet}'"}
    except SheetError:
        raise
    except Exception as e:
        logger.error(f"Error al copiar hoja: {e}")
        raise SheetError(f"Error al copiar hoja: {str(e)}")

def delete_sheet(filepath: str, sheet_name: str) -> Dict[str, Any]:
    """
    Elimina una hoja del libro.
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la hoja a eliminar
        
    Returns:
        Diccionario con el resultado de la operación
        
    Raises:
        SheetError: Si ocurre un error al eliminar la hoja
    """
    try:
        # Abrir el libro
        wb = open_workbook(filepath, read_only=False)
        
        # Verificar que la hoja existe
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"La hoja '{sheet_name}' no existe")
            
        # Verificar que no es la única hoja
        if len(wb.sheetnames) == 1:
            raise SheetError("No se puede eliminar la única hoja del libro")
            
        # Eliminar hoja
        del wb[sheet_name]
        
        # Guardar y cerrar
        wb.save(filepath)
        wb.close()
        
        return {"message": f"Hoja '{sheet_name}' eliminada"}
    except SheetError:
        raise
    except Exception as e:
        logger.error(f"Error al eliminar hoja: {e}")
        raise SheetError(f"Error al eliminar hoja: {str(e)}")

def rename_sheet(filepath: str, old_name: str, new_name: str) -> Dict[str, Any]:
    """
    Renombra una hoja.
    
    Args:
        filepath: Ruta al archivo Excel
        old_name: Nombre actual de la hoja
        new_name: Nuevo nombre para la hoja
        
    Returns:
        Diccionario con el resultado de la operación
        
    Raises:
        SheetError: Si ocurre un error al renombrar la hoja
    """
    try:
        # Abrir el libro
        wb = open_workbook(filepath, read_only=False)
        
        # Verificar que la hoja actual existe
        if old_name not in wb.sheetnames:
            raise SheetError(f"La hoja '{old_name}' no existe")
            
        # Verificar que el nuevo nombre no existe
        if new_name in wb.sheetnames:
            raise SheetError(f"La hoja '{new_name}' ya existe")
            
        # Renombrar hoja
        sheet = wb[old_name]
        sheet.title = new_name
        
        # Guardar y cerrar
        wb.save(filepath)
        wb.close()
        
        return {"message": f"Hoja renombrada de '{old_name}' a '{new_name}'"}
    except SheetError:
        raise
    except Exception as e:
        logger.error(f"Error al renombrar hoja: {e}")
        raise SheetError(f"Error al renombrar hoja: {str(e)}")

def merge_range(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> Dict[str, Any]:
    """
    Combina un rango de celdas.
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        start_cell: Celda inicial
        end_cell: Celda final
        
    Returns:
        Diccionario con el resultado de la operación
        
    Raises:
        SheetError: Si ocurre un error al combinar las celdas
    """
    try:
        # Abrir el libro
        wb = open_workbook(filepath, read_only=False)
        
        # Verificar que la hoja existe
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"La hoja '{sheet_name}' no existe")
            
        ws = wb[sheet_name]
        
        # Analizar rango
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
        except ValueError as e:
            raise SheetError(f"Formato de rango inválido: {str(e)}")
            
        # Verificar que ambas celdas están especificadas
        if end_row is None or end_col is None:
            raise SheetError("Es necesario especificar celda inicial y final para combinar")
            
        # Crear cadena de rango
        range_str = f"{start_cell}:{end_cell}"
        
        # Combinar celdas
        ws.merge_cells(range_str)
        
        # Guardar y cerrar
        wb.save(filepath)
        wb.close()
        
        return {"message": f"Celdas combinadas: {range_str}"}
    except SheetError:
        raise
    except Exception as e:
        logger.error(f"Error al combinar celdas: {e}")
        raise SheetError(f"Error al combinar celdas: {str(e)}")

def unmerge_range(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> Dict[str, Any]:
    """
    Descombina un rango de celdas.
    
    Args:
        filepath: Ruta al archivo Excel
        sheet_name: Nombre de la hoja
        start_cell: Celda inicial
        end_cell: Celda final
        
    Returns:
        Diccionario con el resultado de la operación
        
    Raises:
        SheetError: Si ocurre un error al descombinar las celdas
    """
    try:
        # Abrir el libro
        wb = open_workbook(filepath, read_only=False)
        
        # Verificar que la hoja existe
        if sheet_name not in wb.sheetnames:
            raise SheetError(f"La hoja '{sheet_name}' no existe")
            
        ws = wb[sheet_name]
        
        # Crear cadena de rango
        range_str = f"{start_cell}:{end_cell}"
        
        # Descombinar celdas
        ws.unmerge_cells(range_str)
        
        # Guardar y cerrar
        wb.save(filepath)
        wb.close()
        
        return {"message": f"Celdas descombinadas: {range_str}"}
    except SheetError:
        raise
    except Exception as e:
        logger.error(f"Error al descombinar celdas: {e}")
        raise SheetError(f"Error al descombinar celdas: {str(e)}")