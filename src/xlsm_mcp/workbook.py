"""
Módulo para gestionar libros de Excel con macros (.xlsm).
"""

import os
import logging
from typing import Dict, List, Any, Optional, Union
from pathlib import Path
import openpyxl
from openpyxl.workbook import Workbook

from xlsm_mcp.exceptions import WorkbookError

logger = logging.getLogger("xlsm-mcp")

def create_workbook(filepath: Union[str, Path], with_macros: bool = True) -> None:
    """
    Crea un nuevo libro de Excel, opcionalmente con macros habilitadas.
    
    Args:
        filepath: Ruta donde guardar el archivo
        with_macros: Si es True, crea un archivo .xlsm con macros habilitadas
    
    Raises:
        WorkbookError: Si ocurre un error al crear el libro
    """
    try:
        # Convertir a objeto Path
        file_path = Path(filepath)
        
        # Crear un nuevo libro
        wb = Workbook()
        
        # Asegurar que la extensión sea correcta según el tipo de libro
        if with_macros and file_path.suffix.lower() != '.xlsm':
            file_path = file_path.with_suffix('.xlsm')
        elif not with_macros and file_path.suffix.lower() != '.xlsx':
            file_path = file_path.with_suffix('.xlsx')
        
        # Crear directorio padre si no existe
        file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Guardar el libro
        wb.save(str(file_path))
        logger.info(f"Libro creado correctamente en {file_path}")
    except Exception as e:
        logger.error(f"Error al crear libro: {e}")
        raise WorkbookError(f"No se pudo crear el libro: {str(e)}")

def get_workbook_info(filepath: Union[str, Path], include_macros: bool = False) -> Dict[str, Any]:
    """
    Obtiene información general sobre un libro de Excel.
    
    Args:
        filepath: Ruta al archivo Excel
        include_macros: Si es True, incluye información sobre macros
    
    Returns:
        Diccionario con información del libro
        
    Raises:
        WorkbookError: Si ocurre un error al abrir el libro o extraer información
    """
    try:
        # Convertir a objeto Path
        file_path = Path(filepath)
        
        # Verificar que el archivo existe
        if not file_path.exists():
            raise WorkbookError(f"El archivo {file_path} no existe")
        
        # Abrir el libro
        wb = openpyxl.load_workbook(str(file_path), read_only=True, keep_vba=True)
        
        # Recopilar información
        info = {
            "filepath": str(file_path.absolute()),
            "filename": file_path.name,
            "file_size": file_path.stat().st_size,
            "last_modified": file_path.stat().st_mtime,
            "has_macros": wb.vba_archive is not None,
            "sheet_names": wb.sheetnames,
            "active_sheet": wb.active.title,
            "properties": {
                "title": wb.properties.title or "",
                "subject": wb.properties.subject or "",
                "creator": wb.properties.creator or "",
                "description": wb.properties.description or "",
                "keywords": wb.properties.keywords or "",
                "category": wb.properties.category or ""
            }
        }
        
        # Si se solicita, agrega información de macros
        if include_macros and wb.vba_archive is not None:
            info["macros"] = {
                "vba_project": True,
                # Aquí puedes agregar más detalles si los necesitas
            }
        else:
            info["macros"] = None
        
        # Cerrar el libro
        wb.close()
        
        return info
    except Exception as e:
        logger.error(f"Error al obtener información del libro: {e}")
        raise WorkbookError(f"No se pudo obtener información del libro: {str(e)}")

def open_workbook(filepath: Union[str, Path], read_only: bool = False) -> Workbook:
    """
    Abre un libro de Excel existente.
    
    Args:
        filepath: Ruta al archivo Excel
        read_only: Si es True, abre el libro en modo solo lectura
        
    Returns:
        Objeto Workbook de openpyxl
        
    Raises:
        WorkbookError: Si ocurre un error al abrir el libro
    """
    try:
        # Convertir a objeto Path
        file_path = Path(filepath)
        
        # Verificar que el archivo existe
        if not file_path.exists():
            raise WorkbookError(f"El archivo {file_path} no existe")
        
        # Determinar si el archivo tiene macros
        has_macros = file_path.suffix.lower() == '.xlsm'
        
        # Abrir el libro
        wb = openpyxl.load_workbook(
            str(file_path), 
            read_only=read_only, 
            keep_vba=has_macros
        )
        
        return wb
    except Exception as e:
        logger.error(f"Error al abrir libro: {e}")
        raise WorkbookError(f"No se pudo abrir el libro: {str(e)}")